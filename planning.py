import streamlit as st
import pandas as pd
import random
import io
from itertools import cycle
from datetime import timedelta, datetime 
import locale 
from openpyxl import load_workbook 
from openpyxl.worksheet.protection import SheetProtection

# Tentative de définition de la locale en français
try:
    locale.setlocale(locale.LC_TIME, 'fr_FR.UTF-8') 
except locale.Error:
    pass  # Si la locale française n'est pas disponible, on passe

# Fonction pour obtenir la date en toutes lettres en français 
def format_date(date):
    try:
        return date.strftime('%A %d %B %Y')
    except:
        # Si la locale française n'est pas disponible, on le fait manuellement
        jours_semaine = ['lundi', 'mardi', 'mercredi', 'jeudi', 'vendredi', 'samedi', 'dimanche']
        mois = ['janvier', 'février', 'mars', 'avril', 'mai', 'juin', 'juillet', 'août', 'septembre', 'octobre', 'novembre', 'décembre']
        return f"{jours_semaine[date.weekday()]} {date.day} {mois[date.month - 1]} {date.year}"

# Titre de l'application
st.title("Planificateur de Horaires de Travail en Équipes")

# Entrée pour les plages horaires
st.header("Configurer les Plages Horaires") 
time_slots = st.text_area("Entrez les plages horaires (une par ligne)", "8:00-12:00\n12:00-16:00\n16:00-20:00")
time_slots = time_slots.split('\n')

# Entrée pour les membres de l'équipe
st.header("Configurer les Membres de l'Équipe") 
team_members = st.text_area("Entrez les noms des membres de l'équipe (un par ligne)", "Alice\nBob\nCharlie\nDavid") 
team_members = team_members.split('\n')

# Sélectionner une plage de dates
st.header("Sélectionner une Plage de Dates") 
date_range = st.date_input("Sélectionnez la plage de dates", [pd.Timestamp('today'), pd.Timestamp('today') + timedelta(days=6)])

# Option pour inclure ou exclure les week-ends 
include_weekends = st.checkbox("Inclure les week-ends", value=True)

# Marquer les absences et spécifier les périodes d'absence multiples 
st.header("Marquer les Absences et Périodes Multiples") 
absences = {}

for member in team_members:
    if st.checkbox(f"{member} est absent", key=f"{member}_absent"):
        periods = []
        period_count = st.number_input(f"Nombre de périodes d'absence pour {member}", min_value=1, value=1, key=f"{member}_period_count")
        for i in range(period_count):
            absence_period = st.date_input(f"Période d'absence {i+1} pour {member}", [pd.Timestamp('today'), pd.Timestamp('today')], key=f"{member}_absence_period_{i}")
            periods.append(absence_period)
        absences[member] = periods

# Bouton pour générer le planning
if st.button("Générer le Planning"):

    # Filtrer les membres disponibles pour chaque jour
    def is_member_available(member, current_date):
        if member not in absences:
            return True
        for period in absences[member]:
            start_absence, end_absence = period
            if start_absence <= current_date <= end_absence:
                return False
        return True

    # Initialiser le planning
    all_schedules = []
    # Initialiser le compteur
    assignment_count = {member: 0 for member in team_members}

    # Générer le planning pour chaque jour de la plage de dates
    start_date, end_date = date_range
    current_date = start_date

    while current_date <= end_date:
        # Vérifier si le jour actuel est un week-end et s'il doit être exclu
        if not include_weekends and current_date.weekday() >= 5:
            current_date += timedelta(days=1)
            continue

        # Filtrer les membres disponibles pour la date actuelle
        available_members = [member for member in team_members if is_member_available(member, current_date)]

        if not available_members:
            st.error(f"Aucun membre disponible pour le {current_date}.")
        else:
            # Créer un cycle des membres disponibles pour garantir l'équité
            random.shuffle(available_members)
            member_cycle = cycle(available_members)
            # Initialiser le planning pour la date actuelle
            schedule = pd.DataFrame(index=time_slots, columns=team_members)
            for time_slot in time_slots:
                assigned_member = next(member_cycle)
                for member in team_members:
                    if member == assigned_member:
                        schedule.loc[time_slot, member] = 'Présent'
                        assignment_count[member] += 1  # Incrémenter le compteur
                    else:
                        schedule.loc[time_slot, member] = 'Absent' if not is_member_available(member, current_date) else 'Libre'
            # Ajouter le planning au dictionnaire de tous les plannings
            full_date = format_date(current_date)  # Utilisation de la fonction de formatage de la date
            schedule.insert(0, 'Date', full_date)
            schedule.index.name = 'Horaires'
            all_schedules.append(schedule)

        # Passer à la date suivante
        current_date += timedelta(days=1)

    # Fonction pour styliser le DataFrame
    def highlight_status(val):
        if val == 'Présent':
            color = 'green'
        elif val == 'Absent':
            color = 'red'
        else:
            color = 'white'
        return f'background-color: {color}'

    # Afficher le planning stylisé pour chaque jour
    st.header("Planning Généré")
    for schedule in all_schedules:
        date = schedule['Date'].iloc[0]
        st.subheader(date)
        styled_schedule = schedule.style.applymap(highlight_status, subset=team_members)
        st.dataframe(styled_schedule)

    # Afficher le compteur
    st.header("Compteur de Présence")
    counter_df = pd.DataFrame.from_dict(assignment_count, orient='index', columns=['Nombre de Présences'])
    st.dataframe(counter_df)

    # Convertir les plannings en fichier Excel
    @st.cache_data
    def convert_df_to_excel(schedules, password):
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            row_offset = 0  # Initial row offset
            for schedule in schedules:
                schedule.to_excel(writer, startrow=row_offset, sheet_name='Planning', index=True)
                row_offset += len(schedule) + 2  # Increment by the number of rows in the schedule plus two for spacing
            workbook = writer.book
            worksheet = writer.sheets['Planning']
            # Appliquer le style directement dans Excel
            format1 = workbook.add_format({'bg_color': 'green', 'font_color': 'white'})
            format2 = workbook.add_format({'bg_color': 'red', 'font_color': 'white'})
            row_offset = 1  # Offset to start after the header
            for schedule in schedules:
                for row_num, row_data in enumerate(schedule.itertuples(), start=row_offset):
                    for col_num, cell_data in enumerate(row_data, start=0):  # Inclure l'index
                        if cell_data == 'Présent':
                            worksheet.write(row_num, col_num, cell_data, format1)
                        elif cell_data == 'Absent':
                            worksheet.write(row_num, col_num, cell_data, format2)
                row_offset += len(schedule) + 2  # Increment by the number of rows in the schedule plus two for spacing
        output.seek(0)

        # Charger le fichier Excel généré et le protéger par un mot de passe
        wb = load_workbook(filename=output)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            ws.protection.set_password(password)
        output_protected = io.BytesIO()
        wb.save(output_protected)
        output_protected.seek(0)

        return output_protected

    # Demander un mot de passe pour protéger le fichier Excel
    password = st.text_input("Entrez un mot de passe pour protéger le fichier Excel", type="password")

    # Convertir les plannings stylisés en fichier Excel téléchargeable
    if password:
        excel_data = convert_df_to_excel(all_schedules, password)
        # Bouton de téléchargement du fichier Excel
        st.download_button(label='Télécharger le planning', data=excel_data, file_name='planning_protégé.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    else:
        st.warning("Veuillez entrer un mot de passe pour protéger le fichier Excel.")
